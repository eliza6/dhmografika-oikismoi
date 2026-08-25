#!/usr/bin/env python3
"""
Build a SQLite database of Greek settlement (οικισμός) demographics from
ELSTAT 2021 census tables:

  - Γ01 (C01): permanent population by SEX x AGE GROUP (5-year bins), down to settlement level (8)
  - Γ04 (C04): permanent population by SEX x EDUCATION LEVEL, down to settlement level (8)
  - Β02 (B02): permanent population by AGE GROUP (6 coarse bins) x EDUCATION LEVEL,
               only down to MUNICIPALITY level (5)

Because Β02 stops at municipality level, the AGE x EDUCATION table per settlement is
ESTIMATED using Iterative Proportional Fitting (IPF): for each settlement we take the
municipality's real AGE x EDUCATION table (Β02) as a seed matrix (it supplies the shape
/ correlation between age and education), then rescale it so its row sums match the
settlement's REAL age-group distribution (from Γ01) and its column sums match the
settlement's REAL education distribution (from Γ04, summed over sex) -- both genuine
settlement-level marginals, not estimates. IPF alternately rescales rows and columns
until both match simultaneously. This is strictly better than a single-pass
proportional split against only one marginal (the previous approach here), because the
result is now consistent with *two* independent, real settlement-level facts at once,
not just one.

Geographic code structure (ELSTAT LAU coding): each level adds 2 digits to the code
of its parent, e.g. settlement code '1110101010101' (13 digits) belongs to
municipality '1110101' (7 digits) = code[:7].
"""
import sqlite3
import struct
import unicodedata
import urllib.request
import urllib.error
import http.client
import time
import openpyxl
import numpy as np
from pathlib import Path

# --------------------------------------------------------------------------
# Source files are downloaded on demand from ELSTAT (www.statistics.gr) and
# cached locally in DATA_DIR. Each run first checks whether the file already
# exists on disk; only missing files are downloaded.
# --------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DB_PATH = BASE_DIR / "settlements.db"

_ELSTAT_URL_TEMPLATE = (
    "https://www.statistics.gr/el/statistics"
    "?p_p_id=documents_WAR_publicationsportlet_INSTANCE_VBZOni0vs5VJ"
    "&p_p_lifecycle=2&p_p_state=normal&p_p_mode=view"
    "&p_p_cacheability=cacheLevelPage"
    "&p_p_col_id=column-2&p_p_col_count=4&p_p_col_pos=2"
    "&_documents_WAR_publicationsportlet_INSTANCE_VBZOni0vs5VJ_javax.faces.resource=document"
    "&_documents_WAR_publicationsportlet_INSTANCE_VBZOni0vs5VJ_ln=downloadResources"
    "&_documents_WAR_publicationsportlet_INSTANCE_VBZOni0vs5VJ_documentID={doc_id}"
    "&_documents_WAR_publicationsportlet_INSTANCE_VBZOni0vs5VJ_locale=el"
)

# local_filename -> ELSTAT documentID
SOURCE_FILES = {
    "gamma01_sex_age_settlements.xlsx": "568507",         # Γ01: φύλο x ηλικία, ανά οικισμό
    "gamma04_sex_education_settlements.xlsx": "568504",   # Γ04: φύλο x εκπαίδευση, ανά οικισμό
    "beta02_age_education_municipalities.xlsx": "568511", # Β02: ηλικία x εκπαίδευση, ανά δήμο
    "beta01_employment_age_municipalities.xlsx": "532688",     # Β01: απασχόληση x ηλικία, ανά δήμο
    "beta02_employment_gender_municipalities.xlsx": "532739",  # Β02: απασχόληση x φύλο, ανά δήμο
    "beta03_employment_education_municipalities.xlsx": "532738",  # Β03: απασχόληση x εκπαίδευση, ανά δήμο
    "gamma13_car_household_dimotiki_koinotita.xlsx": "532737",  # Γ13: αυτοκίνητα x νοικοκυριό, ανά δημοτική κοινότητα
}


def _download_once(url: str, timeout: int = 120) -> bytes:
    """Fetch the URL, reading in chunks so a slow/unstable connection doesn't
    trigger an IncompleteRead on the first attempt."""
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
            ),
            "Accept": "*/*",
            "Referer": "https://www.statistics.gr/el/statistics",
            "Connection": "close",
        },
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        chunks = []
        while True:
            chunk = resp.read(262144)  # 256KB at a time
            if not chunk:
                break
            chunks.append(chunk)
        return b"".join(chunks)


def ensure_downloaded(filename: str, doc_id: str, force: bool = False,
                       attempts: int = 4) -> Path:
    """
    Return the local path to `filename`, downloading it from ELSTAT first if
    it isn't already cached in DATA_DIR (or if force=True). Retries a few
    times with backoff since statistics.gr occasionally cuts the connection
    mid-transfer on larger files (IncompleteRead).
    """
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    dest = DATA_DIR / filename

    if dest.exists() and not force:
        if dest.stat().st_size > 0:
            print(f"[cache] {filename} already downloaded, skipping.")
            return dest
        print(f"[cache] {filename} exists but is empty, re-downloading.")

    url = _ELSTAT_URL_TEMPLATE.format(doc_id=doc_id)
    print(f"[download] Fetching {filename} from statistics.gr (documentID={doc_id}) ...")

    last_error = None
    for attempt in range(1, attempts + 1):
        try:
            data = _download_once(url)
            break
        except (urllib.error.URLError, urllib.error.HTTPError,
                http.client.IncompleteRead, ConnectionError, TimeoutError) as e:
            last_error = e
            print(f"[download] Attempt {attempt}/{attempts} failed ({e}); retrying...")
            time.sleep(2 * attempt)
    else:
        raise RuntimeError(
            f"Αποτυχία λήψης του αρχείου '{filename}' από {url} μετά από "
            f"{attempts} προσπάθειες.\n"
            f"Τελευταίο σφάλμα: {last_error}\n"
            f"Κατέβασέ το χειροκίνητα από τον παραπάνω σύνδεσμο και αποθήκευσέ το "
            f"ως: {dest}"
        ) from last_error

    # Basic sanity check: a real .xlsx is a zip archive, starts with 'PK'
    if not data[:2] == b"PK":
        raise RuntimeError(
            f"Το αρχείο που κατέβηκε για '{filename}' δεν μοιάζει με έγκυρο .xlsx "
            f"(πιθανόν η ΕΛΣΤΑΤ επέστρεψε σελίδα σφάλματος αντί για το αρχείο). "
            f"Δοκίμασε να το κατεβάσεις χειροκίνητα από:\n{url}\n"
            f"και να το αποθηκεύσεις ως: {dest}"
        )

    dest.write_bytes(data)
    print(f"[download] Saved {filename} ({len(data):,} bytes) -> {dest}")
    return dest


def get_source_paths(force: bool = False):
    """Ensure all source files are present locally (downloading any that
    are missing) and return a dict of filename -> local Path."""
    paths = {}
    for filename, doc_id in SOURCE_FILES.items():
        paths[filename] = ensure_downloaded(filename, doc_id, force=force)
    return paths

SETTLEMENT_LEVEL = 8
MUNICIPALITY_LEVEL = 5

AGE_5Y_BINS = [" 0-4", " 5-9", "10 -14", "15-19", "20-24", "25-29", "30-34",
               "35-39", "40-44", "45-49", "50-54", "55-59", "60-64", "65-69",
               "70-74", "75+"]

# Mapping from the 16 five-year bins (Γ01) to the 6 coarse bins used in Β02
COARSE_AGE_GROUPS = ["0-14", "15-29", "30-44", "45-59", "60-74", "75+"]
BIN_TO_COARSE = {
    " 0-4": "0-14", " 5-9": "0-14", "10 -14": "0-14",
    "15-19": "15-29", "20-24": "15-29", "25-29": "15-29",
    "30-34": "30-44", "35-39": "30-44", "40-44": "30-44",
    "45-49": "45-59", "50-54": "45-59", "55-59": "45-59",
    "60-64": "60-74", "65-69": "60-74", "70-74": "60-74",
    "75+": "75+",
}

EDU_LEVELS = [
    "Διδακτορικό / Μεταπτυχιακό / Πτυχίο Πανεπιστημίου-Πολυτεχνείου, ΑΤΕΙ, ΑΣΠΑΙΤΕ, "
    "Ανώτερων Επαγγελματικών Σχολών και ισότιμων σχολών",
    "Πτυχίο μεταδευτεροβάθμιας εκπαίδευσης (ΙΕΚ, Κολέγια κλπ.)",
    "Απολυτήριο Λυκείου (Γενικού, Επαγγελματικού, Εκκλησιαστικού κ.λπ. ή εξαταξίου Γυμνασίου)",
    "Πτυχίο Επαγγελματικών Σχολών / Απολυτήριο τριτάξιου Γυμνασίου",
    "Απολυτήριο Δημοτικού",
    "Εγκατέλειψε ή δε φοίτησε στο  Δημοτικό, αλλά γνωρίζει γραφή και ανάγνωση / "
    "Ολοκλήρωσε την προσχολική αγωγή / Δεν γνωρίζει γραφή και ανάγνωση",
    "Μη κατατασσόμενοι (άτομα γεννηθέντα μετά την 1/1/2016)",
]
EDU_LEVELS_SHORT = [
    "Διδακτορικό/Μεταπτυχιακό/Πανεπιστήμιο/ΑΤΕΙ",
    "Μεταδευτεροβάθμια (ΙΕΚ, Κολέγια)",
    "Απολυτήριο Λυκείου",
    "Πτυχίο Επαγγελματικής Σχολής / Γυμνάσιο",
    "Απολυτήριο Δημοτικού",
    "Χωρίς απολυτήριο Δημοτικού (γνώση γραφής-ανάγνωσης)",
    "Μη κατατασσόμενοι (γεννηθέντες μετά 1/1/2016)",
]

# Employment status categories. ELSTAT publishes slightly different levels of
# detail for "οικονομικά μη ενεργοί" across the age-group table (5 sub-cats:
# Μαθητές, Συνταξιούχοι, Εισοδηματίες, Οικιακά, Λοιποί) vs. the gender and
# education tables (3 sub-cats: Μαθητές, Συνταξιούχοι, Λοιποί). We use the
# common denominator across all three so "employment" is one consistent
# dimension everywhere: Εισοδηματίες + Οικιακά fold into "Λοιποί μη ενεργοί"
# for the age-group table specifically.
EMP_CATS = [
    "Απασχολούμενοι",
    "Άνεργοι",
    "Μαθητές-σπουδαστές",
    "Συνταξιούχοι",
    "Λοιποί μη ενεργοί",
]

GENDER_LABELS = ["Άρρενες", "Θήλεις"]

# Number-of-cars-per-household buckets (Γ13)
CAR_CATS = ["0 αυτοκίνητα", "1 αυτοκίνητο", "2 αυτοκίνητα", "3+ αυτοκίνητα"]


def strip_accents(s: str) -> str:
    nfkd = unicodedata.normalize("NFD", s)
    return "".join(c for c in nfkd if unicodedata.category(c) != "Mn")


def normalize(s: str) -> str:
    """Lowercase, accent-stripped, single-spaced -- used for search matching."""
    s = strip_accents(s).lower()
    s = s.replace("ς", "σ")
    return " ".join(s.split())


def split_name_article(raw: str):
    """
    ELSTAT settlement names come as 'Name,article' e.g. 'Κομοτηνή,η',
    sometimes with parenthetical notes e.g. 'Μεγάλη Άδα,η (Ποθαία)'.
    Returns (clean_name, article, display_name).
    """
    raw = raw.strip()
    if "," in raw:
        name, article = raw.split(",", 1)
        name = name.strip()
        article = article.strip()
    else:
        name, article = raw, ""
    display = f"{article} {name}".strip() if article and "(" not in article else name
    # keep it simple/robust: if article has extra notes, just show "Name (άρθρο: X)"
    if article and any(ch in article for ch in "()"):
        display = name
    return name, article, display


def read_rows(path, sheet_name, min_row=6):
    wb = openpyxl.load_workbook(path, data_only=True)
    # ELSTAT files have exactly one sheet each, but its exact name can differ
    # slightly between what we saw locally and what the live download serves
    # (naming is not always 100% stable), so fall back to "whichever sheet is
    # actually there" rather than failing on an exact name mismatch.
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=min_row, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue
        yield row


def ipf(seed, row_margin, col_margin, max_iter=100, tol=1e-6):
    """
    Iterative Proportional Fitting.

    Adjusts `seed` (a rows x cols matrix giving the *shape*/correlation
    between the two variables) so that its row sums match `row_margin` and
    its column sums match `col_margin` exactly (both are assumed to already
    agree on the grand total). Converges in a handful of iterations for
    small matrices like ours (6 age groups x 7 education levels).

    A tiny epsilon is added to every cell first so no cell is permanently
    locked at zero by the seed alone (e.g. a municipality that happens to
    have 0 people in one age x education combination) -- it can still pick
    up population if the settlement's real margins require it.
    """
    n_rows, n_cols = len(seed), len(seed[0])
    mat = [[seed[i][j] + 1e-6 for j in range(n_cols)] for i in range(n_rows)]

    for _ in range(max_iter):
        # --- fit rows to row_margin ---
        for i in range(n_rows):
            s = sum(mat[i])
            target = row_margin[i]
            if s > 0:
                factor = target / s
                mat[i] = [v * factor for v in mat[i]]
            elif target > 0:
                mat[i] = [target / n_cols] * n_cols

        # --- fit columns to col_margin ---
        col_sums = [sum(mat[i][j] for i in range(n_rows)) for j in range(n_cols)]
        for j in range(n_cols):
            s = col_sums[j]
            target = col_margin[j]
            if s > 0:
                factor = target / s
                for i in range(n_rows):
                    mat[i][j] *= factor
            elif target > 0:
                for i in range(n_rows):
                    mat[i][j] = target / n_rows

        # Convergence must be checked AFTER both steps of this iteration --
        # fitting columns perturbs the row sums set moments earlier, so the
        # only valid check is against the matrix's actual current state.
        row_err = max(abs(sum(mat[i]) - row_margin[i]) for i in range(n_rows))
        col_sums_now = [sum(mat[i][j] for i in range(n_rows)) for j in range(n_cols)]
        col_err = max(abs(col_sums_now[j] - col_margin[j]) for j in range(n_cols))
        if row_err < tol and col_err < tol:
            break

    return mat


def scale_rows(seed_rows, real_totals, categories):
    """
    Single-margin proportional fit (used for the employment tables, where we
    only have ONE real settlement-level total to constrain against -- e.g.
    real age totals, but no real settlement-level employment totals since
    ELSTAT never publishes employment status below municipality level).

    seed_rows: dict[category] -> municipality-level employment breakdown (list)
    real_totals: dict[category] -> real settlement-level total for that category
    Returns dict[category] -> rescaled employment breakdown (list), each row
    scaled independently so it sums to real_totals[category].
    """
    out = {}
    for cat in categories:
        seed = seed_rows.get(cat) or [0] * len(EMP_CATS)
        s = sum(seed)
        target = real_totals.get(cat, 0) if isinstance(real_totals, dict) else 0
        if s > 0:
            factor = target / s
            out[cat] = [v * factor for v in seed]
        elif target > 0:
            out[cat] = [target / len(seed)] * len(seed)
        else:
            out[cat] = [0] * len(seed)
    return out


def rake_4d(dims_shape, constraints, max_iter=60, tol=1e-6):
    """
    Generalized IPF ("raking") for a 4D array against multiple pairwise (2D)
    marginal constraints. Standard technique for population synthesis when
    you know several 2-way tables but not the full N-way joint distribution:
    starting from a uniform prior, repeatedly rescale the array so that each
    known 2-way marginal (summing over the other two dimensions) matches its
    target, cycling through all constraints until every one is satisfied
    simultaneously (or max_iter is reached). This is exactly the "4D IPF full
    constrained" step in the pop_synthesis diagram.

    NOTE ON CONVERGENCE: because gender (and to a lesser extent the other
    dimensions) is constrained by MORE THAN ONE target table here (e.g.
    gender totals appear in the gender x age table from Γ01, the gender x
    education table from Γ04, AND the gender x employment estimate), and
    ELSTAT independently rounds/perturbs each published table for
    statistical disclosure control, these targets don't always agree with
    each other exactly (typically within ~0.5-1%). Raking converges to the
    best achievable compromise across all constraints, not to zero error --
    this plateaus after ~40-60 iterations rather than continuing to improve,
    which is expected and not a bug.

    constraints: list of (axes, target) where axes=(i,j) with i<j are the two
    dimension indices this target constrains, and target is a numpy array of
    shape (dims_shape[i], dims_shape[j]).
    """
    arr = np.ones(dims_shape, dtype=np.float64)
    n = len(dims_shape)
    for _ in range(max_iter):
        for axes, target in constraints:
            other_axes = tuple(k for k in range(n) if k not in axes)
            current = arr.sum(axis=other_axes)
            factor = np.ones_like(target, dtype=np.float64)
            pos = current > 1e-9
            factor[pos] = target[pos] / current[pos]
            factor[target <= 1e-9] = 0.0
            full_shape = [1] * n
            full_shape[axes[0]] = dims_shape[axes[0]]
            full_shape[axes[1]] = dims_shape[axes[1]]
            arr = arr * factor.reshape(full_shape)

        # Convergence must be checked AFTER a full pass over every constraint
        # -- checking mid-pass (right after fitting one constraint) is stale,
        # since the next constraint's rescaling immediately perturbs it again.
        max_err = 0.0
        for axes, target in constraints:
            other_axes = tuple(k for k in range(n) if k not in axes)
            current = arr.sum(axis=other_axes)
            if current.size:
                max_err = max(max_err, float(np.max(np.abs(current - target))))
        if max_err < tol:
            break
    return arr


def build(force_download: bool = False):
    global F_SEX_AGE, F_SEX_EDU, F_AGE_EDU_MUNI
    global F_EMP_AGE, F_EMP_GENDER, F_EMP_EDU, F_CAR_HOUSEHOLD
    paths = get_source_paths(force=force_download)
    F_SEX_AGE = paths["gamma01_sex_age_settlements.xlsx"]
    F_SEX_EDU = paths["gamma04_sex_education_settlements.xlsx"]
    F_AGE_EDU_MUNI = paths["beta02_age_education_municipalities.xlsx"]
    F_EMP_AGE = paths["beta01_employment_age_municipalities.xlsx"]
    F_EMP_GENDER = paths["beta02_employment_gender_municipalities.xlsx"]
    F_EMP_EDU = paths["beta03_employment_education_municipalities.xlsx"]
    F_CAR_HOUSEHOLD = paths["gamma13_car_household_dimotiki_koinotita.xlsx"]

    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.executescript("""
    PRAGMA journal_mode=WAL;
    DROP TABLE IF EXISTS settlements;
    DROP TABLE IF EXISTS municipalities;
    DROP TABLE IF EXISTS sex_age;
    DROP TABLE IF EXISTS sex_edu;
    DROP TABLE IF EXISTS age_edu;
    DROP TABLE IF EXISTS edu_levels;
    DROP TABLE IF EXISTS emp_cats;
    DROP TABLE IF EXISTS joint4d;
    DROP TABLE IF EXISTS car_household;

    CREATE TABLE municipalities (
        code TEXT PRIMARY KEY,
        name TEXT
    );

    CREATE TABLE settlements (
        code TEXT PRIMARY KEY,
        raw_name TEXT,
        clean_name TEXT,
        article TEXT,
        display_name TEXT,
        search_key TEXT,
        municipality_code TEXT,
        population INTEGER
    );

    CREATE TABLE sex_age (
        settlement_code TEXT,
        sex TEXT,            -- 'Άρρενες' / 'Θήλεις' / 'Σύνολο'
        age_bin TEXT,         -- 5-year bin
        population INTEGER
    );

    CREATE TABLE sex_edu (
        settlement_code TEXT,
        sex TEXT,
        edu_level TEXT,
        population INTEGER
    );

    CREATE TABLE age_edu (
        settlement_code TEXT,
        age_group TEXT,       -- coarse 6-bucket age group
        edu_level TEXT,
        population REAL,
        is_estimated INTEGER
    );

    CREATE TABLE edu_levels (
        idx INTEGER PRIMARY KEY,
        full_label TEXT,
        short_label TEXT
    );

    CREATE TABLE emp_cats (
        idx INTEGER PRIMARY KEY,
        label TEXT
    );

    -- Single-margin IPF estimates (IPF#2/#3/#4 in the pop_synthesis diagram):
    -- each employment breakdown is only constrained by the ONE real
    -- settlement-level margin available for that pair (age, gender, or
    -- education respectively) -- ELSTAT never publishes a real settlement-
    -- level employment total to use as a second constraint.
    CREATE TABLE gender_emp (
        settlement_code TEXT,
        gender TEXT,
        emp_cat TEXT,
        population REAL,
        is_estimated INTEGER
    );

    CREATE TABLE age_emp (
        settlement_code TEXT,
        age_group TEXT,
        emp_cat TEXT,
        population REAL,
        is_estimated INTEGER
    );

    CREATE TABLE edu_emp (
        settlement_code TEXT,
        edu_level TEXT,
        emp_cat TEXT,
        population REAL,
        is_estimated INTEGER
    );

    -- Full 4D joint distribution (gender x employment x age x education) per
    -- settlement, stored as a compact float32 blob (2*5*6*7 = 420 cells,
    -- ~1.7KB) rather than one row per cell (which would be ~5.7M rows across
    -- all settlements). The API unpacks this on request and sums over
    -- whichever dimensions the drill-down UI hasn't fixed yet.
    CREATE TABLE joint4d (
        settlement_code TEXT PRIMARY KEY,
        blob BLOB,
        total REAL
    );

    -- Number of cars per household. NOT published below Δημοτική Κοινότητα
    -- level, so every settlement in the same Δημοτική Κοινότητα shares the
    -- same distribution (real data, no estimation -- just inherited, per
    -- the "settlement follows municipal unit distribution" assumption).
    CREATE TABLE car_household (
        settlement_code TEXT PRIMARY KEY,
        dk_code TEXT,
        dk_name TEXT,
        total_households INTEGER,
        no_car INTEGER,
        one_car INTEGER,
        two_cars INTEGER,
        three_plus_cars INTEGER
    );

    CREATE INDEX idx_settlements_search ON settlements(search_key);
    CREATE INDEX idx_gender_emp_code ON gender_emp(settlement_code);
    CREATE INDEX idx_age_emp_code ON age_emp(settlement_code);
    CREATE INDEX idx_edu_emp_code ON edu_emp(settlement_code);
    CREATE INDEX idx_sex_age_code ON sex_age(settlement_code);
    CREATE INDEX idx_sex_edu_code ON sex_edu(settlement_code);
    CREATE INDEX idx_age_edu_code ON age_edu(settlement_code);
    """)

    for i, (full, short) in enumerate(zip(EDU_LEVELS, EDU_LEVELS_SHORT)):
        cur.execute("INSERT INTO edu_levels VALUES (?,?,?)", (i, full, short))
    for i, label in enumerate(EMP_CATS):
        cur.execute("INSERT INTO emp_cats VALUES (?,?)", (i, label))

    # ---------- Municipalities (from Γ01, level 5) ----------
    muni_rows = []
    for row in read_rows(F_SEX_AGE, "Γ01"):
        if row[0] == MUNICIPALITY_LEVEL:
            muni_rows.append((str(row[1]), row[2]))
    cur.executemany("INSERT OR IGNORE INTO municipalities VALUES (?,?)", muni_rows)
    print(f"Municipalities: {len(muni_rows)}")

    # ---------- Settlements + sex x age (Γ01) ----------
    # Γ01 layout: col0=level, col1=code, col2=name, col3=Σύνολο(Και των δύο φύλων),
    # col4..col19 = 16 age bins (Σύνολο block), col20=Άρρενες Σύνολο, col21..36=age bins,
    # col37=Θήλεις Σύνολο, col38..53=age bins
    settlement_rows = []
    sex_age_rows = []
    # settlement_code -> {coarse_age_group: total_pop_both_sexes}
    settlement_age_totals = {}
    # settlement_code -> {'Άρρενες': tot, 'Θήλεις': tot} -- REAL settlement-level
    # gender marginal (used later as a constraint for the 4D joint estimate).
    settlement_gender_totals = {}
    # settlement_code -> {'Άρρενες': [6 coarse age vals], 'Θήλεις': [...]} -- REAL
    # settlement-level gender x age table (used as a 4D joint constraint).
    settlement_gender_age = {}

    SEX_BLOCKS = [
        ("Σύνολο", 4),    # start col of the 16 age bins for "both sexes"
        ("Άρρενες", 21),
        ("Θήλεις", 38),
    ]

    for row in read_rows(F_SEX_AGE, "Γ01"):
        if row[0] != SETTLEMENT_LEVEL:
            continue
        code = str(row[1])
        raw_name = row[2]
        total_pop = row[3] or 0
        clean, article, display = split_name_article(raw_name)
        muni_code = code[:7]
        search_key = normalize(clean)
        settlement_rows.append((code, raw_name, clean, article, display, search_key,
                                 muni_code, total_pop))

        coarse_totals = {g: 0 for g in COARSE_AGE_GROUPS}
        gender_coarse = {"Άρρενες": {g: 0 for g in COARSE_AGE_GROUPS},
                          "Θήλεις": {g: 0 for g in COARSE_AGE_GROUPS}}
        for sex_label, start_col in SEX_BLOCKS:
            for j, bin_label in enumerate(AGE_5Y_BINS):
                val = row[start_col + j] or 0
                sex_age_rows.append((code, sex_label, bin_label.strip(), val))
                coarse = BIN_TO_COARSE[bin_label]
                if sex_label == "Σύνολο":
                    coarse_totals[coarse] += val
                else:
                    gender_coarse[sex_label][coarse] += val
        settlement_age_totals[code] = coarse_totals
        settlement_gender_age[code] = gender_coarse
        # row[20] / row[37] are the real per-sex grand totals for the settlement
        settlement_gender_totals[code] = {
            "Άρρενες": row[20] or 0,
            "Θήλεις": row[37] or 0,
        }

    cur.executemany(
        "INSERT INTO settlements VALUES (?,?,?,?,?,?,?,?)", settlement_rows)
    cur.executemany(
        "INSERT INTO sex_age VALUES (?,?,?,?)", sex_age_rows)
    print(f"Settlements: {len(settlement_rows)}  sex_age rows: {len(sex_age_rows)}")

    # ---------- sex x education (Γ04) ----------
    # Γ04 layout: col0=level,1=code,2=name,3=Σύνολο total,4..10=7 edu cats (Σύνολο block)
    # col11=Άρρενες total,12..18=edu cats, col19=Θήλεις total,20..26=edu cats
    sex_edu_rows = []
    # settlement_code -> [7 values] for sex='Σύνολο' -- this is a REAL settlement-level
    # education marginal (unlike age_edu, which has no direct settlement-level source).
    settlement_edu_totals = {}
    # settlement_code -> {'Άρρενες': [7 vals], 'Θήλεις': [7 vals]} -- REAL settlement-level
    # gender x education table (used as a 4D joint constraint).
    settlement_gender_edu = {}
    EDU_SEX_BLOCKS = [("Σύνολο", 4), ("Άρρενες", 12), ("Θήλεις", 20)]
    for row in read_rows(F_SEX_EDU, "Γ04"):
        if row[0] != SETTLEMENT_LEVEL:
            continue
        code = str(row[1])
        for sex_label, start_col in EDU_SEX_BLOCKS:
            vals = [row[start_col + k] or 0 for k in range(7)]
            for k, val in enumerate(vals):
                sex_edu_rows.append((code, sex_label, EDU_LEVELS_SHORT[k], val))
            if sex_label == "Σύνολο":
                settlement_edu_totals[code] = vals
            else:
                settlement_gender_edu.setdefault(code, {})[sex_label] = vals
    cur.executemany("INSERT INTO sex_edu VALUES (?,?,?,?)", sex_edu_rows)
    print(f"sex_edu rows: {len(sex_edu_rows)}")

    # ---------- age x education, municipality level (Β02) ----------
    # Β02 layout: level 5 row = municipality total (skip), followed by 6 rows
    # (one per coarse age group), col0=level,1=code,2=age_label,3=total,4..10=7 edu cats
    # NOTE: in Β02 every row for a given municipality (its total row AND its 6
    # age-group rows) repeats the SAME geographic code in column 1 -- the age
    # rows are distinguished only by their label in column 2, not by a new code.
    muni_age_edu = {}  # (muni_code, age_group) -> [7 values]
    for row in read_rows(F_AGE_EDU_MUNI, "Β02"):
        if row[0] != MUNICIPALITY_LEVEL:
            continue
        code = str(row[1])
        label = row[2]
        if label in COARSE_AGE_GROUPS:
            muni_age_edu[(code, label)] = [row[4 + k] or 0 for k in range(7)]

    age_edu_rows = []
    n_rows, n_cols = len(COARSE_AGE_GROUPS), 7
    skipped_no_seed = 0
    settlement_age_edu_mat = {}  # s_code -> 6x7 matrix (age x edu), reused for the 4D step
    for s_code, muni_code in [(s[0], s[6]) for s in settlement_rows]:
        # Row margin: this settlement's real age-group distribution (Γ01).
        coarse_totals = settlement_age_totals.get(s_code, {})
        row_margin = [coarse_totals.get(a, 0) for a in COARSE_AGE_GROUPS]
        # Column margin: this settlement's real education distribution (Γ04).
        col_margin = settlement_edu_totals.get(s_code, [0] * n_cols)

        # Seed matrix: the municipality's real age x education table (Β02) --
        # this supplies the *shape* (correlation between age and education)
        # that isn't observable at settlement level.
        seed = []
        have_seed = True
        for a in COARSE_AGE_GROUPS:
            vals = muni_age_edu.get((muni_code, a))
            if vals is None:
                have_seed = False
                break
            seed.append(list(vals))

        if not have_seed or sum(row_margin) == 0:
            skipped_no_seed += 1
            continue

        mat = ipf(seed, row_margin, col_margin)
        settlement_age_edu_mat[s_code] = mat
        for i, age_group in enumerate(COARSE_AGE_GROUPS):
            for k in range(n_cols):
                age_edu_rows.append(
                    (s_code, age_group, EDU_LEVELS_SHORT[k], mat[i][k], 1))

    cur.executemany("INSERT INTO age_edu VALUES (?,?,?,?,?)", age_edu_rows)
    print(f"age_edu rows (IPF-estimated): {len(age_edu_rows)}  "
          f"(settlements skipped for missing seed/margin: {skipped_no_seed})")

    # ---------- employment x age group, municipality (Β01) ----------
    # col0=level,1=code,2=desc/age,3=Σύνολο,4=Οικ.ενεργοί Σύνολο,5=Απασχολούμενοι,
    # 6=Άνεργοι Σύνολο,7-8=(υποκατηγορίες, αγνοούνται),9=Οικ.μη ενεργοί Σύνολο,
    # 10=Μαθητές,11=Συνταξιούχοι,12=Εισοδηματίες,13=Οικιακά,14=Λοιποί
    muni_emp_age = {}  # (muni_code, age_group) -> [5 vals] in EMP_CATS order
    for row in read_rows(F_EMP_AGE, "Β01"):
        if row[0] != MUNICIPALITY_LEVEL:
            continue
        code, label = str(row[1]), row[2]
        if label in COARSE_AGE_GROUPS:
            other = (row[12] or 0) + (row[13] or 0) + (row[14] or 0)
            muni_emp_age[(code, label)] = [row[5] or 0, row[6] or 0,
                                            row[10] or 0, row[11] or 0, other]
    print(f"employment x age (municipality) pairs: {len(muni_emp_age)}")

    # ---------- employment x gender, municipality (Β02) ----------
    # col0=level,1=code,2=desc/φύλο,3=Σύνολο,4=Οικ.ενεργοί Σύνολο,5=Απασχολούμενοι,
    # 6=Άνεργοι,7=Οικ.μη ενεργοί Σύνολο,8=Μαθητές,9=Συνταξιούχοι,10=Λοιποί
    muni_emp_gender = {}  # (muni_code, gender) -> [5 vals]
    for row in read_rows(F_EMP_GENDER, "Β02"):
        if row[0] != MUNICIPALITY_LEVEL:
            continue
        code, label = str(row[1]), row[2]
        if label in GENDER_LABELS:
            muni_emp_gender[(code, label)] = [row[5] or 0, row[6] or 0,
                                               row[8] or 0, row[9] or 0, row[10] or 0]
    print(f"employment x gender (municipality) pairs: {len(muni_emp_gender)}")

    # ---------- employment x education, municipality (Β03) ----------
    # same column layout as Β02 above; row[2] is the FULL education label text
    # (matches EDU_LEVELS, not the short version) -- normalize whitespace to match.
    edu_label_lookup = {" ".join(lbl.split()): i for i, lbl in enumerate(EDU_LEVELS)}
    muni_emp_edu = {}  # (muni_code, edu_short_label) -> [5 vals]
    for row in read_rows(F_EMP_EDU, "Β03"):
        if row[0] != MUNICIPALITY_LEVEL:
            continue
        code, label = str(row[1]), row[2]
        idx = edu_label_lookup.get(" ".join(str(label).split()))
        if idx is not None:
            muni_emp_edu[(code, EDU_LEVELS_SHORT[idx])] = [
                row[5] or 0, row[6] or 0, row[8] or 0, row[9] or 0, row[10] or 0]
    print(f"employment x education (municipality) pairs: {len(muni_emp_edu)}")

    # ---------- car number x household, Δημοτική Κοινότητα (Γ13, level 7) ----------
    # Real data -- ELSTAT does not publish this below Δημοτική Κοινότητα, so every
    # settlement inherits its parent Δημοτική Κοινότητα's distribution unchanged
    # ("assume settlement follows municipal unit distribution", per the diagram).
    # col0=level,1=code,2=name,3=Σύνολο νοικοκυριών,4=Χωρίς αυτοκίνητο,
    # 5=Με αυτοκίνητο Σύνολο,6=1 αυτοκίνητο,7=2 αυτοκίνητα,8=3+ αυτοκίνητα
    DK_LEVEL = 7
    dk_car = {}  # dk_code -> (name, total, no_car, one, two, three_plus)
    for row in read_rows(F_CAR_HOUSEHOLD, "Γ13"):
        if row[0] != DK_LEVEL:
            continue
        code = str(row[1])
        dk_car[code] = (row[2], row[3] or 0, row[4] or 0,
                         row[6] or 0, row[7] or 0, row[8] or 0)
    print(f"car x household (Δημοτική Κοινότητα) rows: {len(dk_car)}")

    # ---------- 4D joint distribution: gender x age x education x employment ----------
    # For each settlement: fit age x employment, gender x employment, and
    # education x employment (each against the ONE real settlement-level margin
    # available for that pair -- "single constrained", IPF#2/#3/#4 in the
    # diagram), then rake a full 4D array against all six now-known 2-way
    # tables at once (2 real: gender x age, gender x education; 4 estimated:
    # age x education, age x employment, gender x employment,
    # education x employment) -- the "4D IPF full constrained" step.
    joint4d_rows = []
    car_rows = []
    gender_emp_rows = []
    age_emp_rows = []
    edu_emp_rows = []
    skipped_4d = 0
    AGE_IDX = {a: i for i, a in enumerate(COARSE_AGE_GROUPS)}

    # Car ownership is independent of the employment/4D pipeline (it only needs
    # the settlement's parent Δημοτική Κοινότητα), so it's populated in its own
    # pass rather than being skipped whenever 4D inputs happen to be missing.
    for s in settlement_rows:
        s_code = s[0]
        dk_code = s_code[:11]
        dk = dk_car.get(dk_code)
        if dk:
            name, tot, no_car, one, two, three_plus = dk
            car_rows.append((s_code, dk_code, name, tot, no_car, one, two, three_plus))

    for s in settlement_rows:
        s_code, raw_name, clean, article, display, search_key, muni_code, total_pop = s

        age_edu_mat = settlement_age_edu_mat.get(s_code)
        ga = settlement_gender_age.get(s_code)
        ge = settlement_gender_edu.get(s_code)
        seed_age_emp = {a: muni_emp_age.get((muni_code, a)) for a in COARSE_AGE_GROUPS}
        seed_gender_emp = {g: muni_emp_gender.get((muni_code, g)) for g in GENDER_LABELS}
        seed_edu_emp = {e: muni_emp_edu.get((muni_code, e)) for e in EDU_LEVELS_SHORT}

        real_age = settlement_age_totals.get(s_code, {})
        real_gender = settlement_gender_totals.get(s_code, {})
        real_edu = dict(zip(EDU_LEVELS_SHORT, settlement_edu_totals.get(s_code, [0] * 7)))

        # Each of these 3 single-margin estimates is stored whenever ITS OWN
        # municipality seed is available, independent of whether the fuller
        # 4D estimate below can also be computed for this settlement.
        age_emp = None
        if not any(v is None for v in seed_age_emp.values()):
            age_emp = scale_rows(seed_age_emp, real_age, COARSE_AGE_GROUPS)
            for a in COARSE_AGE_GROUPS:
                for k, cat in enumerate(EMP_CATS):
                    age_emp_rows.append((s_code, a, cat, age_emp[a][k], 1))

        gender_emp = None
        if not any(v is None for v in seed_gender_emp.values()):
            gender_emp = scale_rows(seed_gender_emp, real_gender, GENDER_LABELS)
            for gnd in GENDER_LABELS:
                for k, cat in enumerate(EMP_CATS):
                    gender_emp_rows.append((s_code, gnd, cat, gender_emp[gnd][k], 1))

        edu_emp = None
        if not any(v is None for v in seed_edu_emp.values()):
            edu_emp = scale_rows(seed_edu_emp, real_edu, EDU_LEVELS_SHORT)
            for e in EDU_LEVELS_SHORT:
                for k, cat in enumerate(EMP_CATS):
                    edu_emp_rows.append((s_code, e, cat, edu_emp[e][k], 1))

        if (age_edu_mat is None or ga is None or ge is None
                or age_emp is None or gender_emp is None or edu_emp is None):
            skipped_4d += 1
            continue

        gender_age_mat = np.array(
            [[ga[g][a] for a in COARSE_AGE_GROUPS] for g in GENDER_LABELS], dtype=float)
        gender_edu_mat = np.array([ge[g] for g in GENDER_LABELS], dtype=float)
        age_edu_arr = np.array(age_edu_mat, dtype=float)
        age_emp_arr = np.array([age_emp[a] for a in COARSE_AGE_GROUPS], dtype=float)
        gender_emp_arr = np.array([gender_emp[g] for g in GENDER_LABELS], dtype=float)
        edu_emp_arr = np.array([edu_emp[e] for e in EDU_LEVELS_SHORT], dtype=float)

        # internal raking axes: gender=0, age=1, education=2, employment=3
        dims_shape = (2, 6, 7, 5)
        constraints = [
            ((0, 1), gender_age_mat),
            ((0, 2), gender_edu_mat),
            ((1, 2), age_edu_arr),
            ((1, 3), age_emp_arr),
            ((0, 3), gender_emp_arr),
            ((2, 3), edu_emp_arr),
        ]
        arr = rake_4d(dims_shape, constraints)

        # reorder to (gender, employment, age, education) to match the UI's
        # drill-down path order before serializing as a compact blob.
        storage_arr = arr.transpose(0, 3, 1, 2)
        blob = storage_arr.astype("float32").tobytes()
        total = float(storage_arr.sum())
        joint4d_rows.append((s_code, blob, total))

    cur.executemany("INSERT INTO joint4d VALUES (?,?,?)", joint4d_rows)
    cur.executemany("INSERT INTO gender_emp VALUES (?,?,?,?,?)", gender_emp_rows)
    cur.executemany("INSERT INTO age_emp VALUES (?,?,?,?,?)", age_emp_rows)
    cur.executemany("INSERT INTO edu_emp VALUES (?,?,?,?,?)", edu_emp_rows)
    cur.executemany("INSERT INTO car_household VALUES (?,?,?,?,?,?,?,?)", car_rows)
    print(f"joint4d rows (settlements with full 4D estimate): {len(joint4d_rows)}  "
          f"(skipped for missing muni seed/margin: {skipped_4d})")
    print(f"gender_emp rows: {len(gender_emp_rows)}  age_emp rows: {len(age_emp_rows)}  "
          f"edu_emp rows: {len(edu_emp_rows)}")
    print(f"car_household rows: {len(car_rows)}")

    con.commit()
    con.close()
    print("Done ->", DB_PATH)


if __name__ == "__main__":
    import sys
    force = "--force" in sys.argv
    build(force_download=force)
